import openpyxl,os
wb = openpyxl.load_workbook('excel/example.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')
# 取表格数据
# for i in range(1,5):
#     print(i,sheet.cell(row = i,column = 3).value)

# 获取表格宽和高
# print(sheet.max_row,sheet.max_column)

# 从表中取得行和列
# print(tuple(sheet))
# for rowofcellobjects in sheet['A1':'C3']:
#     for cellobj in rowofcellobjects:
#         #sheet['A1':'C3']包含三个元组:每个元组代表一行。
#         #每个元组内部是该行的三个单元格，单元格Cell对象由value属性(单元格内容)、row属性(单元格行数)、column属性(单元格列数)和coordinate(单元格坐标)属性。
#         print(cellobj.coordinate, cellobj.value) 
#     print('--- END OF ROW ---')

